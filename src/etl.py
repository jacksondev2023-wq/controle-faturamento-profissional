
import re
import unicodedata
from datetime import datetime
from pathlib import Path
from typing import Optional, Tuple, List

import pandas as pd
import numpy as np
from openpyxl import load_workbook

MONTHS = {
    1: "Jan", 2: "Fev", 3: "Mar", 4: "Abr", 5: "Mai", 6: "Jun",
    7: "Jul", 8: "Ago", 9: "Set", 10: "Out", 11: "Nov", 12: "Dez"
}

MONTH_NAME_TO_NUM = {
    "JAN": 1, "JANEIRO": 1,
    "FEV": 2, "FEVEREIRO": 2,
    "MAR": 3, "MARCO": 3, "MARÇO": 3,
    "ABR": 4, "ABRIL": 4,
    "MAI": 5, "MAIO": 5,
    "JUN": 6, "JUNHO": 6,
    "JUL": 7, "JULHO": 7,
    "AGO": 8, "AGOSTO": 8,
    "SET": 9, "SETEMBRO": 9,
    "OUT": 10, "OUTUBRO": 10,
    "NOV": 11, "NOVEMBRO": 11,
    "DEZ": 12, "DEZEMBRO": 12,
}

DEFAULT_DEPARA = pd.DataFrame({
    "sigla_origem": [
        "HM - Paraíba", "HM PARAIBA", "HM PB", "Milagres João Pessoa",
        "HR - CG", "HR CG", "Hospital Residencial Campina Grande",
        "AHC - Natal", "AHC NATAL", "Aliança Natal",
        "Natal Home", "NATAL HOME", "Life Home PE", "LIFE HOME PE",
        "HM AM", "Hospital Milagres Amazonas",
        "Hospital Residencial", "Hospital Residencial - CG",
    ],
    "nome_padrao": [
        "MILAGRES HOME CARE - JP", "MILAGRES HOME CARE - JP", "MILAGRES HOME CARE - JP", "MILAGRES HOME CARE - JP",
        "HOSPITAL RESIDENCIAL - CG", "HOSPITAL RESIDENCIAL - CG", "HOSPITAL RESIDENCIAL - CG",
        "ALIANCA HOME CARE - RN", "ALIANCA HOME CARE - RN", "ALIANCA HOME CARE - RN",
        "NATAL HOME CARE - RN", "NATAL HOME CARE - RN", "LIFE HOME CARE - PE", "LIFE HOME CARE - PE",
        "MILAGRES HOME CARE - AM", "MILAGRES HOME CARE - AM",
        "HOSPITAL RESIDENCIAL - CG", "HOSPITAL RESIDENCIAL - CG",
    ]
})

EXTRA_DEPARA_UNIDADES = pd.DataFrame({
    "sigla_origem": [
        "AHC - SE", "ALIANCA HOME CARE - SE",
        "HM - Brasília", "HM - BRASILIA", "MILAGRES HOME CARE - DF",
        "HM - Cuiabá", "HM - CUIABA", "MILAGRES HOME CARE - MT",
        "HM - Fortaleza", "MILAGRES HOME CARE - CE",
        "HM - Goiânia", "HM - GOIANIA", "MILAGRES HOME CARE - GO",
        "HM - Manaus", "MILAGRES HOME CARE - AM",
        "HM - Rondônia", "HM - RONDONIA", "MILAGRES HOME CARE - RO",
        "HR - BA", "SAUDE BAHIA",
        "HR - JP", "HOSPITAL RESIDENCIAL - JP",
        "HR - MACEIO", "LIFE HOME CARE - AL",
        "HR - RECIFE", "LIFE HOME CARE - PE",
        "UNION-RJ", "UNION CARE - RJ",
        "UNION-SP", "UC - SÃO PAULO", "UC - SAO PAULO",
        "NATAL HOME", "Natal Home", "NATAL HOME CARE", "NATAL HOME CARE - RN",
    ],
    "nome_padrao": [
        "ALIANCA HOME CARE - SE", "ALIANCA HOME CARE - SE",
        "MILAGRES HOME CARE - DF", "MILAGRES HOME CARE - DF", "MILAGRES HOME CARE - DF",
        "MILAGRES HOME CARE - MT", "MILAGRES HOME CARE - MT", "MILAGRES HOME CARE - MT",
        "MILAGRES HOME CARE - CE", "MILAGRES HOME CARE - CE",
        "MILAGRES HOME CARE - GO", "MILAGRES HOME CARE - GO", "MILAGRES HOME CARE - GO",
        "MILAGRES HOME CARE - AM", "MILAGRES HOME CARE - AM",
        "MILAGRES HOME CARE - RO", "MILAGRES HOME CARE - RO", "MILAGRES HOME CARE - RO",
        "SAUDE BAHIA", "SAUDE BAHIA",
        "HOSPITAL RESIDENCIAL - JP", "HOSPITAL RESIDENCIAL - JP",
        "LIFE HOME CARE - AL", "LIFE HOME CARE - AL",
        "LIFE HOME CARE - PE", "LIFE HOME CARE - PE",
        "UNION CARE - RJ", "UNION CARE - RJ",
        "UC - SAO PAULO", "UC - SAO PAULO", "UC - SAO PAULO",
        "NATAL HOME CARE", "NATAL HOME CARE", "NATAL HOME CARE", "NATAL HOME CARE",
    ],
})

DEFAULT_DEPARA = pd.concat([DEFAULT_DEPARA, EXTRA_DEPARA_UNIDADES], ignore_index=True)

DEFAULT_OPERADORA_DEPARA = pd.DataFrame({
    "sigla_origem": [
        "GOV SUS", "AFRAFEP SAUDE",
        "AMIL - AM/PB/PE/RN",
        "ASSEFAZ - CE", "ASSEFAZ - JP/CG",
        "CAMED - JP/CG",
        "CAPESESP - AM", "CAPESESP - CE", "CAPESESP - PB", "CAPESESP - RN",
        "CASSI - JP/CG", "CASSI - MT", "CASSI - RN",
        "CAURN - RN",
        "FACHESF - PE/BA",
        "FUNSA AERONAUTICA",
        "FUSEX - AM", "FUSEX - CG", "FUSEX - DF", "FUSEX - JP", "FUSEX - RN", "FUSEX - RO",
        "FUSMA - CE", "FUSMA - JP/CG",
        "GEAP - ALAGOAS", "GEAP - BAHIA", "GEAP - BRASILIA", "GEAP - CUIABA",
        "GEAP - GOIAS", "GEAP - MANAUS", "GEAP - PARAIBA", "GEAP - PERNAMBUCO",
        "GEAP - RIO DE JANEIRO", "GEAP - RIO G DO NORTE", "GEAP - RONDONIA", "GEAP - SERGIPE",
        "HAPVIDA (PROCESSO JUDICIALIZADO)",
        "HUMANA - RN",
        "INAS - DF",
        "MARINHA - RN",
        "POSTAL SAUDE", "POSTAL SAUDE - AM",
        "SAUDE CAIXA - JP/CG",
        "SESAP - RN",
        "TRF - MANAUS",
        "TRFMED- RECIFE",
        "UNIMED JP  - JP/CG", "UNIMED JP - JP/CG",
        "UNIMED RECIFE - PE", "UNIMED RECIFE - JP",
        "UNIMED PATOS  - JP", "UNIMED PATOS - JP",
        "UNIMED NOVA IGUACU",
    ],
    "nome_padrao": [
        "SUS", "AFRAFEP",
        "AMIL",
        "ASSEFAZ", "ASSEFAZ",
        "CAMED",
        "CAPESESP", "CAPESESP", "CAPESESP", "CAPESESP",
        "CASSI", "CASSI", "CASSI",
        "CAURN",
        "FACHESF",
        "FUNSA",
        "FUSEX", "FUSEX", "FUSEX", "FUSEX", "FUSEX", "FUSEX",
        "FUSMA", "FUSMA",
        "GEAP", "GEAP", "GEAP", "GEAP",
        "GEAP", "GEAP", "GEAP", "GEAP",
        "GEAP", "GEAP", "GEAP", "GEAP",
        "HAPVIDA",
        "HUMANA",
        "INAS",
        "MARINHA",
        "POSTAL", "POSTAL",
        "SAUDE CAIXA",
        "SESAP",
        "TRF",
        "TRFMED",
        "UNIMED JP", "UNIMED JP",
        "UNIMED RECIFE", "UNIMED RECIFE",
        "UNI PATOS", "UNI PATOS",
        "UNI NOCA IGUA",
    ],
})

FAT_ALIASES = {
    "unidade": ["UNIDADE", "FILIAL", "EMPRESA"],
    "operadora": ["OPERADORA", "CONVENIO CONSOLIDADO", "CONVÊNIO CONSOLIDADO", "CONVENIO", "CONVÊNIO"],
    "valor": ["VALOR_COBRAR", "VALOR A COBRAR", "Valor a Cobrar", "VALOR FATURADO", "VALOR"],
    "data": ["COMPETENCIA FAT", "DATA_EMISSAO_GUIA", "DATA EMISSAO GUIA", "DATA DE EMISSÃO", "VIGENCIA_DE", "Vigência de:"],
    "nf": ["IDDOC", "ID Doc", "ID DOC", "Nº NF", "NF"],
    "paciente": ["PACIENTE", "Nome do Paciente", "NOME PACIENTE"],
}

CONT_ALIASES = {
    "nf": ["Nº NF", "NF", "NOTA", "NOTA FISCAL"],
    "unidade": ["UNIDADE", "FILIAL", "EMPRESA"],
    "operadora": ["OPERADORA", "CONVENIO", "CONVÊNIO"],
    "bruto": ["VALOR BRUTO", "BRUTO", "VALOR_BRUTO"],
    "liquido": ["VALOR LÍQUIDO", "VALOR LIQUIDO", "LÍQUIDO", "LIQUIDO", "VALOR_LIQUIDO"],
    "data_pago": ["DTA DE PAGO", "DATA DE PAGO", "DTA_PAGO", "PAGO EM", "PAGAMENTO"],
    "mes_recebimento": ["MÊS DE RECEBIMENTO", "MES DE RECEBIMENTO", "MÊS RECEBIMENTO", "MES_RECEBIMENTO", "MÊS PAGAMENTO", "MES PAGAMENTO"],
    "obs": ["OBSERVAÇÕES", "OBSERVACOES", "OBS", "OBSERVAÇÃO", "OBSERVACAO"],
}

def norm_text(value) -> str:
    if pd.isna(value):
        return ""
    value = str(value).strip().upper()
    value = re.sub(r"\s+", " ", value)
    replacements = {
        "Á":"A", "À":"A", "Â":"A", "Ã":"A",
        "É":"E", "Ê":"E",
        "Í":"I",
        "Ó":"O", "Ô":"O", "Õ":"O",
        "Ú":"U", "Ç":"C",
    }
    for old, new in replacements.items():
        value = value.replace(old, new)
    value = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode("ascii")
    return value

def find_col(df: pd.DataFrame, aliases: List[str]) -> Optional[str]:
    lookup = {norm_text(col): col for col in df.columns}
    for alias in aliases:
        key = norm_text(alias)
        if key in lookup:
            return lookup[key]
    for alias in aliases:
        key = norm_text(alias)
        for col_norm, col in lookup.items():
            if key in col_norm or col_norm in key:
                return col
    return None

def parse_money(series: pd.Series) -> pd.Series:
    if pd.api.types.is_numeric_dtype(series):
        return pd.to_numeric(series, errors="coerce").fillna(0.0)
    s = series.astype(str)
    neg = s.str.contains(r"^\s*\(.*\)\s*$", regex=True)
    cleaned = (
        s.str.replace("R$", "", regex=False)
         .str.replace(" ", "", regex=False)
         .str.replace(".", "", regex=False)
         .str.replace(",", ".", regex=False)
         .str.replace("(", "", regex=False)
         .str.replace(")", "", regex=False)
    )
    values = pd.to_numeric(cleaned, errors="coerce").fillna(0.0)
    values[neg] = -values[neg].abs()
    return values

def parse_month_year(value, fallback_year: int = 2026) -> Tuple[Optional[int], Optional[int]]:
    if pd.isna(value):
        return None, None
    if isinstance(value, (pd.Timestamp, datetime)):
        return int(value.month), int(value.year)
    parsed = pd.to_datetime(value, errors="coerce", dayfirst=True)
    if pd.notna(parsed):
        return int(parsed.month), int(parsed.year)
    s = norm_text(value).replace("_", " ").replace("/", " ").replace("-", " ")
    parts = [p for p in re.split(r"\W+", s) if p]
    month = None
    year = None
    for p in parts:
        if p in MONTH_NAME_TO_NUM:
            month = MONTH_NAME_TO_NUM[p]
        elif p.isdigit():
            n = int(p)
            if 1 <= n <= 12 and month is None:
                month = n
            elif n >= 2000:
                year = n
            elif 0 <= n <= 99 and year is None:
                year = 2000 + n
    return month, year or fallback_year

def standardize_unit(series: pd.Series, depara: pd.DataFrame) -> pd.Series:
    mapping = {norm_text(k): norm_text(v) for k, v in zip(depara["sigla_origem"], depara["nome_padrao"])}
    return series.apply(lambda x: mapping.get(norm_text(x), norm_text(x)))

def standardize_operator(unidade_series: pd.Series, operadora_series: pd.Series, depara: Optional[pd.DataFrame] = None) -> pd.Series:
    if depara is None or depara.empty:
        depara = DEFAULT_OPERADORA_DEPARA.copy()
        
    temp = pd.DataFrame({"u": unidade_series.apply(norm_text), "o": operadora_series.apply(norm_text)})
    has_u = "unidade_origem" in depara and not depara["unidade_origem"].isnull().all()
    
    dict_map = {}
    dict_map_general = {}
    
    if has_u:
        for u, o, p in zip(depara["unidade_origem"].fillna(""), depara["sigla_origem"].fillna(""), depara["nome_padrao"].fillna("")):
            u_norm, o_norm, p_norm = norm_text(u), norm_text(o), norm_text(p)
            if u_norm:
                dict_map[(u_norm, o_norm)] = p_norm
            else:
                dict_map_general[o_norm] = p_norm
    else:
        for o, p in zip(depara["sigla_origem"].fillna(""), depara["nome_padrao"].fillna("")):
            dict_map_general[norm_text(o)] = norm_text(p)
            
    def get_padrao(row):
        u_val, o_val = row["u"], row["o"]
        
        # 1. Heurística de Judicialização (IW)
        # Se a operadora é "particular", "judic" ou é o nome da própria unidade (ou contida nela), unificamos.
        if o_val and u_val and (o_val in u_val or u_val in o_val):
            # Cuidado com falsos positivos de nomes curtos, mas nomes de unidades geralmente são longos (ex: LIFE HOME CARE - PE)
            if len(o_val) > 8:
                return "PROCESSO JUDICIALIZADO"
                
        if o_val and ("PARTIC" in o_val or "JUDIC" in o_val):
            return "PROCESSO JUDICIALIZADO"
            
        # Regra solicitada: Hapvida na Life entra como judicializado/particular
        if "HAPVIDA" in o_val and "LIFE" in u_val:
            return "PROCESSO JUDICIALIZADO"
            
        # 2. De-Para Customizado
        if (u_val, o_val) in dict_map:
            return dict_map[(u_val, o_val)]
        if o_val in dict_map_general:
            return dict_map_general[o_val]
            
        # 3. Fallback
        return o_val
        
    return temp.apply(get_padrao, axis=1)

def clean_observacao_fiscal(series: pd.Series) -> pd.Series:
    s = series.fillna("").astype(str)
    return s.where(s.str.upper().str.contains("FILIAL FISCAL|FISCAL", regex=True), "")

def read_first_sheet(path_or_file) -> pd.DataFrame:
    xls = pd.ExcelFile(path_or_file)
    sheet = xls.sheet_names[0]
    df = pd.read_excel(path_or_file, sheet_name=sheet)
    return df

def prepare_faturamento(
    df: pd.DataFrame,
    depara: Optional[pd.DataFrame] = None,
    depara_operadoras: Optional[pd.DataFrame] = None,
    fallback_year: int = 2026,
    origem: str = "",
) -> pd.DataFrame:
    if depara is None:
        depara = DEFAULT_DEPARA.copy()
    if df.empty:
        return df

    unidade = find_col(df, FAT_ALIASES["unidade"])
    operadora = find_col(df, FAT_ALIASES["operadora"])
    valor = find_col(df, FAT_ALIASES["valor"])
    data = find_col(df, FAT_ALIASES["data"])
    nf = find_col(df, FAT_ALIASES["nf"])
    paciente = find_col(df, FAT_ALIASES["paciente"])

    missing = [k for k, v in {"unidade": unidade, "operadora": operadora, "valor": valor, "data": data}.items() if not v]
    if missing:
        raise ValueError(f"Colunas não encontradas no faturamento: {', '.join(missing)}")

    out = pd.DataFrame()
    out["nf"] = df[nf].astype(str) if nf else ""
    out["unidade_original"] = df[unidade]
    out["unidade_padrao"] = standardize_unit(df[unidade], depara)
    out["operadora_original"] = df[operadora]
    out["operadora_padrao"] = standardize_operator(out["unidade_padrao"], df[operadora], depara_operadoras)
    out["paciente"] = df[paciente] if paciente else ""
    out["valor_faturado"] = parse_money(df[valor])
    months = df[data].apply(lambda v: parse_month_year(v, fallback_year))
    out["mes_faturamento"] = months.apply(lambda x: x[0])
    out["ano_faturamento"] = months.apply(lambda x: x[1])
    out["origem_arquivo"] = origem
    out["fonte"] = "FATURAMENTO_RAW"
    return out

def prepare_contabilidade(
    df: pd.DataFrame,
    depara: Optional[pd.DataFrame] = None,
    depara_operadoras: Optional[pd.DataFrame] = None,
    fallback_year: int = 2026,
    origem: str = "",
) -> pd.DataFrame:
    if depara is None:
        depara = DEFAULT_DEPARA.copy()
    if df.empty:
        return df

    unidade = find_col(df, CONT_ALIASES["unidade"])
    operadora = find_col(df, CONT_ALIASES["operadora"])
    bruto = find_col(df, CONT_ALIASES["bruto"])
    liquido = find_col(df, CONT_ALIASES["liquido"])
    mes_col = find_col(df, CONT_ALIASES["mes_recebimento"])
    data_pago = find_col(df, CONT_ALIASES["data_pago"])
    nf = find_col(df, CONT_ALIASES["nf"])
    obs = find_col(df, CONT_ALIASES["obs"])

    missing = [k for k, v in {"unidade": unidade, "operadora": operadora, "bruto": bruto, "liquido": liquido}.items() if not v]
    if missing or not (mes_col or data_pago):
        raise ValueError("Colunas essenciais não encontradas na contabilidade.")

    out = pd.DataFrame()
    out["nf"] = df[nf].astype(str) if nf else ""
    out["unidade_original"] = df[unidade]
    out["unidade_padrao"] = standardize_unit(df[unidade], depara)
    out["operadora_original"] = df[operadora]
    out["operadora_padrao"] = standardize_operator(out["unidade_padrao"], df[operadora], depara_operadoras)
    out["valor_bruto"] = parse_money(df[bruto])
    out["valor_liquido"] = parse_money(df[liquido])
    if data_pago:
        out["data_pago"] = pd.to_datetime(df[data_pago], errors="coerce", dayfirst=True)
    else:
        out["data_pago"] = pd.NaT
    ref = df[mes_col] if mes_col else df[data_pago]
    months = ref.apply(lambda v: parse_month_year(v, fallback_year))
    out["mes_recebimento"] = months.apply(lambda x: x[0])
    out["ano_recebimento"] = months.apply(lambda x: x[1])
    out["observacao_original"] = df[obs] if obs else ""
    out["observacao_fiscal"] = clean_observacao_fiscal(df[obs]) if obs else ""
    out["origem_arquivo"] = origem
    out["fonte"] = "CONTABILIDADE_RAW"
    return out

def prepare_consolidado_historico(df: pd.DataFrame, origem: str = "") -> pd.DataFrame:
    # Transforma relatórios já consolidados em histórico consultável.
    if df.empty:
        return df
    unidade = find_col(df, ["Unidade", "UNIDADE"])
    operadora = find_col(df, ["Operadora", "OPERADORA"])
    obs = find_col(df, ["Observação", "OBSERVACAO", "OBSERVAÇÕES"])
    rows = []

    for _, row in df.iterrows():
        unidade_padrao = norm_text(row.get(unidade, ""))
        operadora_padrao = norm_text(row.get(operadora, ""))
        if not unidade_padrao or not operadora_padrao:
            continue
        obs_val = str(row.get(obs, "")) if obs else ""
        obs_fiscal = obs_val if re.search("FILIAL FISCAL|FISCAL", obs_val.upper()) else ""

        for col in df.columns:
            col_norm = norm_text(col)
            valor = row.get(col, 0)
            try:
                valor_num = float(valor) if pd.notna(valor) else 0.0
            except Exception:
                valor_num = 0.0

            # Ex: Faturado Mar, Rec. Bruto Abr, Rec. Líquido Mai
            mes = None
            for nome, n in MONTH_NAME_TO_NUM.items():
                if f" {nome}" in f" {col_norm}" or col_norm.endswith(nome):
                    mes = n
                    break
            if mes is None:
                continue

            tipo = None
            if "FATURADO" in col_norm:
                tipo = "FATURADO"
            elif "BRUTO" in col_norm and "REC" in col_norm:
                tipo = "RECEBIDO_BRUTO"
            elif ("LIQUIDO" in col_norm or "LÍQUIDO" in col_norm) and "REC" in col_norm:
                tipo = "RECEBIDO_LIQUIDO"
            if tipo:
                rows.append({
                    "unidade_padrao": unidade_padrao,
                    "operadora_padrao": operadora_padrao,
                    "tipo": tipo,
                    "mes": mes,
                    "ano": 2026,
                    "valor": valor_num,
                    "observacao_fiscal": obs_fiscal,
                    "origem_arquivo": origem,
                    "fonte": "CONSOLIDADO_HISTORICO",
                })
    return pd.DataFrame(rows)

DINAMICA_COLUMNS = [
    "linha_origem",
    "unidade_original",
    "unidade_padrao",
    "operadora_original",
    "operadora_padrao",
    "faturado_marco",
    "faturado_abril",
    "rec_bruto_marco",
    "rec_liquido_marco",
    "rec_bruto_abril",
    "rec_liquido_abril",
    "rec_bruto_maio",
    "rec_liquido_maio",
    "alerta_diretoria",
    "sinal_diretoria",
    "observacao",
    "origem_arquivo",
    "atualizado_em",
]

def _cell_number(value) -> float:
    if pd.isna(value):
        return 0.0
    if isinstance(value, str) and not value.strip():
        return 0.0
    try:
        return float(value)
    except Exception:
        return float(parse_money(pd.Series([value])).iloc[0])

def _color_rgb(color) -> str:
    rgb = getattr(color, "rgb", None)
    if not rgb:
        return ""
    return str(rgb).upper()[-6:]

def _is_director_alert_rgb(rgb: str) -> bool:
    if not rgb or len(rgb) != 6:
        return False
    try:
        red, green, blue = int(rgb[0:2], 16), int(rgb[2:4], 16), int(rgb[4:6], 16)
    except ValueError:
        return False
    return red >= 200 and green <= 215 and blue <= 215 and (blue >= 80 or green < 140)

def _cell_has_director_alert(cell) -> bool:
    fill_rgb = _color_rgb(getattr(getattr(cell, "fill", None), "fgColor", None))
    font_rgb = _color_rgb(getattr(getattr(cell, "font", None), "color", None))
    return _is_director_alert_rgb(fill_rgb) or _is_director_alert_rgb(font_rgb)

def _dinamica_column_key(header_value) -> Optional[str]:
    header = norm_text(header_value)
    if not header:
        return None
    if "OBSERV" in header:
        return "observacao"

    month = None
    for month_name, month_num in MONTH_NAME_TO_NUM.items():
        if re.search(rf"(^|[^A-Z]){re.escape(norm_text(month_name))}", header):
            month = month_num
            break
    if month is None:
        return None

    month_slug = {3: "marco", 4: "abril", 5: "maio"}.get(month)
    if not month_slug:
        return None
    if "FATURADO" in header:
        return f"faturado_{month_slug}"
    if "REC" in header and "BRUTO" in header:
        return f"rec_bruto_{month_slug}"
    if "REC" in header and ("LIQUIDO" in header or "LQUIDO" in header):
        return f"rec_liquido_{month_slug}"
    return None

def parse_dinamica_workbook(path_or_file, origem: str = "") -> pd.DataFrame:
    """Lê somente a aba DINAMICA e retorna linhas analíticas de operadora.

    A planilha recebida é uma tabela dinâmica: linhas em negrito são subtotais
    por unidade; linhas sem negrito são operadoras. Os subtotais da planilha são
    ignorados para que o sistema recalcule tudo.
    """
    wb = load_workbook(path_or_file, data_only=True)
    sheet_name = next((name for name in wb.sheetnames if norm_text(name) == "DINAMICA"), None)
    if not sheet_name:
        raise ValueError("A aba DINAMICA não foi encontrada no arquivo.")
    ws = wb[sheet_name]

    unit_candidates: set[str] = set()
    for candidate_sheet in wb.worksheets:
        if candidate_sheet.title == sheet_name:
            continue
        header_lookup = {}
        for col in range(1, candidate_sheet.max_column + 1):
            header_lookup[norm_text(candidate_sheet.cell(1, col).value)] = col
        unit_col = header_lookup.get("UNIDADE")
        operator_col = header_lookup.get("OPERADORA")
        if not unit_col or not operator_col:
            continue
        for row in range(2, candidate_sheet.max_row + 1):
            unit_value = str(candidate_sheet.cell(row, unit_col).value or "").strip()
            operator_value = str(candidate_sheet.cell(row, operator_col).value or "").strip()
            if unit_value and operator_value:
                unit_candidates.add(norm_text(unit_value))

    header_row = None
    for row in range(1, min(ws.max_row, 30) + 1):
        header_label = norm_text(ws.cell(row, 1).value)
        if header_label in {"UNIDADES X OPERADORA", "UNIDADE X OPERADORA", "ROTULOS DE LINHA", "RÓTULOS DE LINHA"}:
            header_row = row
            break
    if header_row is None:
        raise ValueError("Cabeçalho de linhas da aba DINAMICA não encontrado.")

    column_map: dict[str, int] = {}
    for col in range(2, ws.max_column + 1):
        key = _dinamica_column_key(ws.cell(header_row, col).value)
        if key and key in DINAMICA_COLUMNS and key not in column_map:
            column_map[key] = col
    if not any(key.startswith("faturado_") for key in column_map):
        raise ValueError("Nenhuma coluna de faturamento foi identificada na aba DINAMICA.")

    rows = []
    current_unit = ""
    now = datetime.now().strftime("%d/%m/%Y %H:%M")
    source_name = origem or getattr(path_or_file, "name", "") or "DINAMICA"

    for row in range(header_row + 1, ws.max_row + 1):
        label = str(ws.cell(row, 1).value or "").strip()
        if not label:
            continue
        if norm_text(label).startswith("TOTAL"):
            continue

        is_unit_row = norm_text(label) in unit_candidates or bool(ws.cell(row, 1).font and ws.cell(row, 1).font.bold)
        if is_unit_row:
            current_unit = label
            continue
        if not current_unit:
            continue

        values = {
            "faturado_marco": 0.0,
            "faturado_abril": 0.0,
            "rec_bruto_marco": 0.0,
            "rec_liquido_marco": 0.0,
            "rec_bruto_abril": 0.0,
            "rec_liquido_abril": 0.0,
            "rec_bruto_maio": 0.0,
            "rec_liquido_maio": 0.0,
        }
        for key, col in column_map.items():
            if key in values:
                values[key] = _cell_number(ws.cell(row, col).value)
        observacao_col = column_map.get("observacao")
        observacao = str(ws.cell(row, observacao_col).value or "").strip() if observacao_col else ""
        if not any(abs(v) > 0.00001 for v in values.values()) and not observacao:
            continue

        rows.append({
            "linha_origem": int(row),
            "unidade_original": current_unit,
            "unidade_padrao": current_unit,
            "operadora_original": label,
            "operadora_padrao": label,
            **values,
            "observacao": observacao,
            "origem_arquivo": source_name,
            "atualizado_em": now,
        })

    return pd.DataFrame(rows, columns=DINAMICA_COLUMNS)

def parse_dinamica_workbook(path_or_file, origem: str = "") -> pd.DataFrame:
    """Le a aba analitica consolidada e retorna linhas de operadora."""
    wb = load_workbook(path_or_file, data_only=True)
    preferred_names = {"DINAMICA", "RELATORIO", "ANALITICO", "RELATORIO ANALITICO"}
    sheet_name = next((name for name in wb.sheetnames if norm_text(name) in preferred_names), None)
    if not sheet_name:
        for candidate in wb.worksheets:
            for row in range(1, min(candidate.max_row, 30) + 1):
                header_label = norm_text(candidate.cell(row, 1).value)
                if header_label in {"UNIDADES X OPERADORA", "UNIDADE X OPERADORA", "UNIDADE / OPERADORA", "ROTULOS DE LINHA"}:
                    sheet_name = candidate.title
                    break
            if sheet_name:
                break
    if not sheet_name:
        raise ValueError("A aba analitica consolidada nao foi encontrada no arquivo.")
    ws = wb[sheet_name]

    unit_candidates: set[str] = set()
    for candidate_sheet in wb.worksheets:
        if candidate_sheet.title == sheet_name:
            continue
        header_lookup = {}
        for col in range(1, candidate_sheet.max_column + 1):
            header_lookup[norm_text(candidate_sheet.cell(1, col).value)] = col
        unit_col = header_lookup.get("UNIDADE")
        operator_col = header_lookup.get("OPERADORA")
        if not unit_col or not operator_col:
            continue
        for row in range(2, candidate_sheet.max_row + 1):
            unit_value = str(candidate_sheet.cell(row, unit_col).value or "").strip()
            operator_value = str(candidate_sheet.cell(row, operator_col).value or "").strip()
            if unit_value and operator_value:
                unit_candidates.add(norm_text(unit_value))

    header_row = None
    for row in range(1, min(ws.max_row, 30) + 1):
        header_label = norm_text(ws.cell(row, 1).value)
        if header_label in {"UNIDADES X OPERADORA", "UNIDADE X OPERADORA", "UNIDADE / OPERADORA", "ROTULOS DE LINHA"}:
            header_row = row
            break
    if header_row is None:
        raise ValueError("Cabecalho de linhas da aba analitica nao encontrado.")

    column_map: dict[str, int] = {}
    type_col = None
    for col in range(2, ws.max_column + 1):
        if norm_text(ws.cell(header_row, col).value) == "TIPO":
            type_col = col
            continue
        key = _dinamica_column_key(ws.cell(header_row, col).value)
        if key and key in DINAMICA_COLUMNS and key not in column_map:
            column_map[key] = col
    if not any(key.startswith("faturado_") for key in column_map):
        raise ValueError("Nenhuma coluna de faturamento foi identificada na aba analitica.")

    rows = []
    current_unit = ""
    now = datetime.now().strftime("%d/%m/%Y %H:%M")
    source_name = origem or getattr(path_or_file, "name", "") or "DINAMICA"

    for row in range(header_row + 1, ws.max_row + 1):
        label = str(ws.cell(row, 1).value or "").strip()
        if not label:
            continue
        if norm_text(label).startswith("TOTAL"):
            continue

        type_label = norm_text(ws.cell(row, type_col).value) if type_col else ""
        if type_label == "UNIDADE":
            current_unit = label
            continue
        if type_label and type_label != "OPERADORA":
            continue

        is_unit_row = not type_col and (
            norm_text(label) in unit_candidates
            or bool(ws.cell(row, 1).font and ws.cell(row, 1).font.bold)
        )
        if is_unit_row:
            current_unit = label
            continue
        if not current_unit:
            continue

        values = {
            "faturado_marco": 0.0,
            "faturado_abril": 0.0,
            "rec_bruto_marco": 0.0,
            "rec_liquido_marco": 0.0,
            "rec_bruto_abril": 0.0,
            "rec_liquido_abril": 0.0,
            "rec_bruto_maio": 0.0,
            "rec_liquido_maio": 0.0,
        }
        for key, col in column_map.items():
            if key in values:
                values[key] = _cell_number(ws.cell(row, col).value)
        observacao_col = column_map.get("observacao")
        observacao = str(ws.cell(row, observacao_col).value or "").strip() if observacao_col else ""
        if not any(abs(v) > 0.00001 for v in values.values()) and not observacao:
            continue

        rows.append({
            "linha_origem": int(row),
            "unidade_original": current_unit,
            "unidade_padrao": current_unit,
            "operadora_original": label,
            "operadora_padrao": label,
            **values,
            "alerta_diretoria": int(_cell_has_director_alert(ws.cell(row, 1))),
            "sinal_diretoria": "vermelho" if _cell_has_director_alert(ws.cell(row, 1)) else "",
            "observacao": observacao,
            "origem_arquivo": source_name,
            "atualizado_em": now,
        })

    return pd.DataFrame(rows, columns=DINAMICA_COLUMNS)

def dinamica_to_raw_tables(base: pd.DataFrame, year: int = 2026, origem: str = "DINAMICA") -> tuple[pd.DataFrame, pd.DataFrame]:
    if base is None or base.empty:
        fat_cols = [
            "nf", "unidade_original", "unidade_padrao", "operadora_original", "operadora_padrao",
            "paciente", "valor_faturado", "mes_faturamento", "ano_faturamento", "origem_arquivo", "fonte",
        ]
        cont_cols = [
            "nf", "unidade_original", "unidade_padrao", "operadora_original", "operadora_padrao",
            "valor_bruto", "valor_liquido", "data_pago", "mes_recebimento", "ano_recebimento",
            "observacao_original", "observacao_fiscal", "origem_arquivo", "fonte",
        ]
        return pd.DataFrame(columns=fat_cols), pd.DataFrame(columns=cont_cols)

    fat_rows = []
    cont_rows = []
    for idx, row in base.reset_index(drop=True).iterrows():
        unit_original = str(row.get("unidade_original", "") or "").strip()
        unit_standard = str(row.get("unidade_padrao", "") or unit_original).strip()
        op_original = str(row.get("operadora_original", "") or "").strip()
        op_standard = str(row.get("operadora_padrao", "") or op_original).strip()
        source = str(row.get("origem_arquivo", "") or origem)
        obs = str(row.get("observacao", "") or "").strip()
        line = int(row.get("linha_origem", idx + 1) or idx + 1)

        for month, value_col in [(3, "faturado_marco"), (4, "faturado_abril")]:
            value = _cell_number(row.get(value_col, 0))
            if abs(value) <= 0.00001:
                continue
            fat_rows.append({
                "nf": f"DIN-{line}-FAT-{month:02d}",
                "unidade_original": unit_original,
                "unidade_padrao": unit_standard,
                "operadora_original": op_original,
                "operadora_padrao": op_standard,
                "paciente": "",
                "valor_faturado": value,
                "mes_faturamento": int(month),
                "ano_faturamento": int(year),
                "origem_arquivo": source,
                "fonte": "DINAMICA_IMPORTADA",
            })

        for month, bruto_col, liquido_col in [
            (3, "rec_bruto_marco", "rec_liquido_marco"),
            (4, "rec_bruto_abril", "rec_liquido_abril"),
            (5, "rec_bruto_maio", "rec_liquido_maio"),
        ]:
            bruto = _cell_number(row.get(bruto_col, 0))
            liquido = _cell_number(row.get(liquido_col, 0))
            obs_for_month = obs if month == 3 else ""
            if abs(bruto) <= 0.00001 and abs(liquido) <= 0.00001 and not obs_for_month:
                continue
            cont_rows.append({
                "nf": f"DIN-{line}-REC-{month:02d}",
                "unidade_original": unit_original,
                "unidade_padrao": unit_standard,
                "operadora_original": op_original,
                "operadora_padrao": op_standard,
                "valor_bruto": bruto,
                "valor_liquido": liquido,
                "data_pago": pd.NaT,
                "mes_recebimento": int(month),
                "ano_recebimento": int(year),
                "observacao_original": obs_for_month,
                "observacao_fiscal": obs_for_month,
                "origem_arquivo": source,
                "fonte": "DINAMICA_IMPORTADA",
            })

    return pd.DataFrame(fat_rows), pd.DataFrame(cont_rows)

def build_consolidado(faturamento: pd.DataFrame, contabilidade: pd.DataFrame, fat_months: list[int], rec_months: list[int], year: int = 2026) -> pd.DataFrame:
    fat = faturamento.copy()
    cont = contabilidade.copy()
    
    # === PROCV POR SEMELHANÇA ===
    # A contabilidade dita a regra de nome de operadora. Se houver discrepância por sufixos (ex: AMIL vs AMIL - PB),
    # o faturamento herdará o nome idêntico da contabilidade automaticamente.
    if not cont.empty and not fat.empty:
        cont_ops = cont[['unidade_padrao', 'operadora_padrao']].drop_duplicates()
        
        import re
        def super_norm(x):
            return re.sub(r'[^a-zA-Z0-9]', '', str(x).upper())
            
        cont_ops['u_norm'] = cont_ops['unidade_padrao'].apply(super_norm)
        
        def match_operator(row):
            u_fat_raw = str(row['unidade_padrao']).strip()
            o_fat = str(row['operadora_padrao']).strip()
            if not o_fat:
                return o_fat
                
            u_fat_norm = super_norm(u_fat_raw)
            ops_na_cont = cont_ops[cont_ops['u_norm'] == u_fat_norm]['operadora_padrao'].tolist()
            
            if o_fat in ops_na_cont:
                return o_fat
                
            # Procura por substring (Faturamento contido na Contabilidade ou vice-versa)
            for o_cont in ops_na_cont:
                o_c_str = str(o_cont).strip()
                if not o_c_str:
                    continue
                # Se o nome do faturamento for longo o suficiente e estiver contido no da contabilidade
                if len(o_fat) >= 4 and o_fat in o_c_str:
                    return o_cont
                # Ou se a contabilidade estiver contida no faturamento
                if len(o_c_str) >= 4 and o_c_str in o_fat:
                    return o_cont
            return o_fat
            
        fat['operadora_padrao'] = fat.apply(match_operator, axis=1)
    # === FIM PROCV ===

    base = pd.DataFrame(columns=["unidade_padrao", "operadora_padrao"])
    fat_cols = []
    for m in fat_months:
        col = f"fat_{m}"
        fat_cols.append(col)
        if fat.empty:
            fat_g = pd.DataFrame(columns=["unidade_padrao", "operadora_padrao", col])
        else:
            f = fat[(fat["ano_faturamento"] == year) & (fat["mes_faturamento"] == m)].copy()
            fat_g = f.groupby(["unidade_padrao", "operadora_padrao"], as_index=False)["valor_faturado"].sum()
            fat_g = fat_g.rename(columns={"valor_faturado": col})
        base = base.merge(fat_g, on=["unidade_padrao", "operadora_padrao"], how="outer") if not base.empty else fat_g

    if base.empty:
        base = pd.DataFrame(columns=["unidade_padrao", "operadora_padrao"] + fat_cols)

    for m in rec_months:
        if cont.empty:
            rec_g = pd.DataFrame(columns=["unidade_padrao", "operadora_padrao", f"rec_bruto_{m}", f"rec_liquido_{m}", f"obs_fiscal_{m}"])
        else:
            c = cont[(cont["ano_recebimento"] == year) & (cont["mes_recebimento"] == m)].copy()
            rec_g = c.groupby(["unidade_padrao", "operadora_padrao"], as_index=False).agg(
                **{
                    f"rec_bruto_{m}": ("valor_bruto", "sum"),
                    f"rec_liquido_{m}": ("valor_liquido", "sum"),
                    f"obs_fiscal_{m}": ("observacao_fiscal", lambda x: " | ".join([v for v in x.astype(str).unique() if v and v != "nan"]))
                }
            )
        base = base.merge(rec_g, on=["unidade_padrao", "operadora_padrao"], how="outer")

    value_cols = [c for c in base.columns if c.startswith("fat_") or c.startswith("rec_bruto_") or c.startswith("rec_liquido_") or c == "faturado"]
    base[value_cols] = base[value_cols].fillna(0)
    base["faturado"] = base[fat_cols].sum(axis=1) if fat_cols else 0
    bruto_cols = [c for c in base.columns if c.startswith("rec_bruto_")]
    liq_cols = [c for c in base.columns if c.startswith("rec_liquido_")]
    obs_cols = [c for c in base.columns if c.startswith("obs_fiscal_")]
    base["total_recebido_bruto"] = base[bruto_cols].sum(axis=1) if bruto_cols else 0
    base["total_recebido_liquido"] = base[liq_cols].sum(axis=1) if liq_cols else 0
    base["diferenca_pendente"] = base["faturado"] - base["total_recebido_bruto"]
    base["perc_recebido_total"] = np.where(base["faturado"] > 0, base["total_recebido_bruto"] / base["faturado"], 0)
    base["observacao_fiscal"] = base[obs_cols].fillna("").agg(lambda r: " | ".join([v for v in r if v]), axis=1) if obs_cols else ""
    return base.sort_values(["diferenca_pendente"], ascending=False)
